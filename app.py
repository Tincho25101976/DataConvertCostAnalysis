import ctypes
import os
import re
import sys
import threading
from typing import Any, Callable, Final

import tkinter as tk
from tkinter import ttk, filedialog
from unittest import result 
from matplotlib import table
from numpy import maximum
from pandas import DataFrame
from polars import first
from traitlets import HasTraits
import ttkbootstrap as tb
from ttkbootstrap.dialogs import Messagebox, Querybox

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import range_boundaries
import pandas as pd

from setting.Setting import Setting
from helper.HelperControl import HelperControl, ControlStyleType, ScrolledCheckboxList, ScrolledProgressList
from process.DataTypeProcess import DataTypeDataTable


class ImportProcess(tb.Window):
    __TITLE_DEFAULT:Final[str] = 'Cost analysis process'
    #__CAPTION_DICT_HEADERS = 'headers'
    #__CAPTION_DICT_ROWS = 'rows'

    def __init__(self, themename:str='superhero'):
        super().__init__(themename=themename)

        self.__appSetDarkTitleBar()

        self.title(self.__TITLE_DEFAULT)
        self.state("zoomed")
        self.geometry("1440x880")
        self.protocol("WM_DELETE_WINDOW", self.onCloseApplication)

        self.__appClearConsole()
        appIcon = self.__appSetIcon()
        if appIcon: self.iconphoto(True, appIcon)

        _style:tb.Style | None = tb.Style.get_instance()
        if not _style:
            return
        self.settingApp: Setting = Setting()
        self.hp: HelperControl = HelperControl(self.settingApp, _style, ControlStyleType.THEME)

        self.__setEnvirotmeVariables()
        self.__setPanels(self.hp)

    #region set app
    def __appSetDarkTitleBar(self):
        try:
            self.update()
            hwnd = ctypes.windll.user32.GetParent(self.winfo_id())

            DWMWA_USE_IMMERSIVE_DARK_MODE = 20
            value = ctypes.c_int(1)

            ctypes.windll.dwmapi.DwmSetWindowAttribute(
                hwnd, DWMWA_USE_IMMERSIVE_DARK_MODE, ctypes.byref(value), ctypes.sizeof(value)
            )
        except Exception as e:
            pass
        
    def __appClearConsole(self):
        sys.stdout.write("\033[H\033[2J")
        sys.stdout.flush()

    def __appSetIcon(self)->tk.PhotoImage | None:
        iconPath = self.__getPath(["app_icon.png"])
        if os.path.exists(iconPath):
            return tk.PhotoImage(file=iconPath)
        return None

    def onCloseApplication(self):
        self.__workbookSecureClose(self.sourceWorkbook)
        self.__workbookSecureClose(self.sourceWorkbookMeta)
        
        #if hasattr(self, '_excelData'): self._excelData = {}
        if hasattr(self, '_dfDataList'): self._dfDataList = None
        if hasattr(self, '_dfDataListIsProcess'): self._dfDataListIsProcess = None

        self.destroy()
        self.quit()
        sys.exit(0)
    #endregion

    #region path
    # ---------------------------------------------------------------------------------------
    # Path:
    # ---------------------------------------------------------------------------------------
    def __getPath(self, folder)->str:
        supportDir:str = os.path.dirname(os.path.abspath(__file__))
        samplesFolder:str = os.path.join(supportDir, *folder)
        return samplesFolder
    #endregion

    #region helper
    def __setValuesComboBox(self, _ctr:tb.Combobox, data:list[str] | None = None):
        if not _ctr: return
        _isInSelf = _ctr in self.__dict__.values()
        if not _isInSelf: return
        _ctr.set('')
        _ctr.configure(values=[])
        if data is None: return
        _ctr.configure(values=data)
        _ctr.set(data[0])
    def __workbookSecureClose(self, wb:Workbook | None):
        if not wb: return
        _isInSelf = wb in self.__dict__.values()
        if not _isInSelf: return
        wb.close()
        wb = None
    def __isInSelf(self, _ctr)->bool:
        if not _ctr: return False
        result:bool = _ctr in self.__dict__.values()
        return result
    def __isActive(self, _ctr)->bool:
        if not _ctr: return False
        return self.__isInSelf(_ctr) and _ctr is not None
    #endregion

    #region panels
    def __setEnvirotmeVariables(self):
        self.sourceFilePath:str | None = os.path.join(self.__getPath(['source']), 'IndexCostAnalysis.xlsx')
        self.sourceWorkbook:Workbook | None = None
        self.sourceWorkbookMeta: Workbook | None = None

        self._excelSheetActive: tk.StringVar = tk.StringVar()
        self._excelTableActive:tk.StringVar = tk.StringVar()
        self._excelProgressFile:tk.DoubleVar = tk.DoubleVar(value=0)
        self._dfDataList:list[DataTypeDataTable] | None = None
        self._dfDataListIsProcess:list[DataTypeDataTable] | None = None

        self._structTableNameActive:tk.StringVar = tk.StringVar()
        self._structSheetNameActive:tk.StringVar = tk.StringVar()
        self._structTableAlias:tk.StringVar = tk.StringVar()
        self._structTableIsProcess:tk.BooleanVar = tk.BooleanVar(value=False)

    def __setPanels(self, hp:HelperControl):
        if hp is None: return

        #region main layout
        # Panel set:
        _frmLeftPanel = hp.ControlFrameLeftPanel(self)
        _notebookLeft = hp.ControlNotebook(_frmLeftPanel)

        _frmTabLeftOne = hp.ControlFrameForColumnEnd(_notebookLeft)
        #_frmTabTwo = hp.ControlFrameForColumnEnd(_notebookLeft)

        _notebookLeft.add(_frmTabLeftOne, text="Excel file")
        #_notebookLeft.add(_frmTabTwo, text="Image summary")

        hp.ControlSeparator(self)
        # Panel get:
        _frmRightPanel = hp.ControlFrameRightPanel(self)
        _notebookRight = hp.ControlNotebook(_frmRightPanel)
        _frmTabRightOne = hp.ControlFrameForColumnEnd(_notebookRight)
        _frmTabRightTwo = hp.ControlFrameForColumnEnd(_notebookRight)

        _notebookRight.add(_frmTabRightOne, text="Viewer")
        _notebookRight.add(_frmTabRightTwo, text="Struct")
        #endregion

        #region left
        _frmPathExcelFile = hp.ControlLabelFrameForRow(_frmTabLeftOne, "Workbook")

        _frmRow0:tb.Frame = hp.ControlFrameForRow(_frmPathExcelFile)
        self._cboSheet:tb.Combobox = hp.ControlComboBoxWithLabelWidth(_frmRow0, 'Sheets', 
                                            _command=self.__wbLoadTableBySheet,
                                            _textVariable=self._excelSheetActive,
                                            _width=18)
        self._cboTable:tb.Combobox = hp.ControlComboBoxWithLabelWidth(_frmRow0, 'Table',
                                            _textVariable=self._excelTableActive,
                                            _expand=True)

        _frmRow1:tb.Frame = hp.ControlFrameForRow(_frmPathExcelFile)
        self._cmdOpen = hp.ControlButtonLEFT(_frmRow1, 'Open', _command=self.__wbReadStructFile)
        self._cmdRead = hp.ControlButtonLEFT(_frmRow1, 'Read', _command=self.__wbLoadAsync, _state=False)
        self._cmdDataTable = hp.ControlButtonRIGHT(_frmRow1, 'Data table', _command=self.__onTableSelected, _state=False)
        
        _frmRow2:tb.Frame = hp.ControlFrameForRow(_frmPathExcelFile)
        self._lblStatus:tb.Label = hp.ControlLabelInfoLEFT(_frmRow2, '...')

        _frmRow3:tb.Frame = hp.ControlFrameForRow(_frmPathExcelFile)
        self.progress_bar = hp.ControlProgressBar(_frmRow3, _variable=self._excelProgressFile)
        self.lstProgress:ScrolledProgressList = hp.ControlScrolledProgressList(
                                                        _frmRow3, 
                                                        "Sheets load progress",
                                                        self.settingApp.GetTables())

        _frmSheetsProcess:tb.Labelframe = hp.ControlLabelFrameForRow(_frmTabLeftOne, "Sheets of process")
        self._lstChkSheetIsProcess:ScrolledCheckboxList = hp.ControlScrolledCheckboxList(
                                                        _frmSheetsProcess, 
                                                        None,
                                                        self.settingApp.GetProcessableTables())

        
        
        _frmNullPanel = hp.ControlFrameForRowEnd(_frmTabLeftOne)
        #endregion

        #region right
        #------------------------------------------------------------------------------------
        # Viewer
        #------------------------------------------------------------------------------------
        _frmViewer:tb.Labelframe = hp.ControlLabelFrameForRowEnd(_frmTabRightTwo, 'Data')
        self._twViewer:tb.Treeview = hp.ControlTreeviewClear(_frmViewer)

        #------------------------------------------------------------------------------------
        # Struct
        #------------------------------------------------------------------------------------
        _frmStruct:tb.LabelFrame = hp.ControlLabelFrameForRowEnd(_frmTabRightOne, "Metadata")
        #hp.SetHelperStyle(_frmStruct, '#4567aa')

        _frmRowStruct01:tb.Frame = hp.ControlFrameForRow(_frmStruct)
        self._cboStructTable:tb.Combobox = hp.ControlComboBoxWithLabelWidth(_frmRowStruct01, 'Tables', 
                                            _textVariable=self._structTableNameActive,
                                            _command=self.__wsStructMakeStruct,
                                            _width=45)
        self._txtStructSheet:tb.Entry = hp.ControlEntryWithLabelWidth(_frmRowStruct01, 'Sheet', 
                                            _textVariable=self._structSheetNameActive,
                                            _status=False)
    
        self._txtTableAlias:tb.Entry = hp.ControlEntryWithLabelWidth(_frmRowStruct01, 'Alias', 
                                            _textVariable=self._structTableAlias)
        
        self._cmdStructSave:tb.Button = hp.ControlButtonRIGHT(_frmRowStruct01, 'Save struct', 
                                            _command=self.__wsStructSave)

        
        _frmRowStruct02:tb.Frame = hp.ControlFrameForRow(_frmStruct)
        self._chkStructIsProcess:tb.Checkbutton = hp.ControlCheckbox(_frmRowStruct02, 'Is process table', 
                                            _variable=self._structTableIsProcess)


        _frmRowStruct03:tb.Frame = hp.ControlFrameForRowEnd(_frmStruct)
        self._lstChkStructColumns:ScrolledCheckboxList = hp.ControlScrolledCheckboxList(_frmRowStruct03, 'Columns of table')
        #hp.SetHelperStyle(_frmRowStruct03)
        
        #endregion
        #------------------------------------------------------
        # set:
    #endregion

    #region workbook actions
    def __wbLoadAsync(self):
        self._cmdRead.config(state=tk.DISABLED)
        self.progress_bar['value'] = 0
        self._dfDataListIsProcess = []
        self._dfDataList = []

        if self.sourceFilePath is None: return

        threading.Thread(
            target=self.__wbReadFile, args=(self.sourceFilePath, self.__wbProgressLoad), daemon=True
        ).start()

    def __wbReadStructFile(self):
        try:
            self.__setValuesComboBox(self._cboStructTable)
            filePath = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
            if not filePath:
                    return
            self.sourceFilePath = filePath
            self.sourceWorkbookMeta = load_workbook(filePath, data_only=True, read_only=False)

            _sheetWithTable = []
            _tablesOfBook:list[str] = []
            for sheetName in self.sourceWorkbookMeta.sheetnames:
                ws:Worksheet = self.sourceWorkbookMeta[sheetName]
                if ws.tables and ws.sheet_state == 'visible':
                    _sheetWithTable.append(sheetName)
                    _tables = [str(s) for s in ws.tables.keys()]
                    if _tables:
                        _tablesOfBook.extend(_tables)

            if not _tablesOfBook:
                self.after(0, self.__wbIssueLoadFile, "Tables not found (ListObjects).")
                return

            #self.__setValuesComboBox(self._cboStructTable, _tablesOfBook)

            if self.sourceWorkbookMeta:
                self.after(0, self.__setValuesComboBox, self._cboStructTable, _tablesOfBook)
                self.after(0, self._cmdRead.config(state=tk.NORMAL))

        except Exception as e:
            self.after(0, self.__wbIssueLoadFile, str(e))
    
    def __wsStructMakeStruct(self):
        tableName:str = self._structTableNameActive.get()
        if not tableName or self.sourceWorkbookMeta is None: return
        wb:Workbook = self.sourceWorkbookMeta 
        _columns:list[str] = []
        if not wb: return
        for sheet_name in wb.sheetnames:
            ws:Worksheet = wb[sheet_name]
            if tableName in ws.tables.keys():
                table = ws.tables[tableName]
                _t = [str(col.name) for col in table.tableColumns]
                _columns.extend(_t)
                _sheet:str | None = self.settingApp.GetSheetName(tableName)
                _alias:str = self.settingApp.GetTableAlias(tableName)
                _isProcess:bool = self.settingApp.IsTableProcessable(tableName)

                self._structSheetNameActive.set(_sheet if not _sheet is None else sheet_name)
                self._structTableAlias.set(_alias)
                self._structTableIsProcess.set(_isProcess)

        if not _columns: return
        if not self.__isInSelf(self._lstChkSheetIsProcess): return
        result:list[tuple[str, bool]] = []
        _columnsSet:list[str] | None = self.settingApp.GetColumnsByTable(tableName)
        if not _columnsSet:
            result = [(s, True) for s in _columns]
        else:
            result = [(s, (s in _columnsSet)) for s in _columns]
            
        self._lstChkStructColumns.SetItems(result)

    def __wsStructSave(self):
        tableName:str = self._structTableNameActive.get()
        tableIsProcess:bool = self._structTableIsProcess.get()
        tableColumns:list[str] = self._lstChkStructColumns.GetCheckedItems()
        sheetName:str = self._structSheetNameActive.get()
        tableAlias:str = self._structTableAlias.get()
        if tableName and tableColumns:
            self.settingApp.SaveOrUpdateTableConfig(tableName, tableColumns, tableIsProcess,
                                sheetName, tableAlias)

    def __wbReadFile(self, filePath: str, progressCallback: Callable[[float, str, str, float], None] | None = None):
        try:
            if not self.__isInSelf(self.sourceWorkbookMeta): return
            if not self.__isInSelf(self.settingApp): return
            if self.sourceWorkbookMeta is None: return
            if self.settingApp is None: return

            '''
            _tablesInSetting: list[str] = self.settingApp.GetTables()
            if not _tablesInSetting: return

            _targetTables: dict[str, dict] = {}
            _setTablesInSetting = set(_tablesInSetting) # Optimización con set

            for sheetName in self.sourceWorkbookMeta.sheetnames:
                ws = self.sourceWorkbookMeta[sheetName]
                if ws.tables and ws.sheet_state == 'visible':
                    for tableName, tableObj in ws.tables.items():
                        if tableName in _setTablesInSetting:
                            _targetTables[str(tableName)] = {
                                'sheet': sheetName,
                                'ref': str(tableObj)
                            }
            '''

            _tablesInSetting: list[tuple[str, str]] = self.settingApp.GetTablesWithSheet()
            if not _tablesInSetting: return

            _targetTables: dict[str, dict] = {}
            #_setTablesInSetting = _tablesInSetting

            for table, sheet in _tablesInSetting:
                ws = self.sourceWorkbookMeta[sheet]
                if ws.tables and ws.sheet_state == 'visible':
                    for tableName, tableObj in ws.tables.items():
                        if tableName == table:
                            _targetTables[str(tableName)] = {
                                'sheet': sheet,
                                'ref': str(tableObj)
                            }

            if not _targetTables:
                self.after(0, self.__wbIssueLoadFile, "Tables not found (ListObjects).")
                return

            self.sourceWorkbook = load_workbook(filePath, data_only=True, read_only=True)

            _rowCount: int = 0
            for tbl_info in _targetTables.values():
                min_c, min_r, max_c, max_r = range_boundaries(tbl_info['ref'])
                if all(s for s in [max_r, min_r] if isinstance(s, int)):
                    max_r_v:int = int(max_r) if isinstance(max_r, int) else 0
                    min_r_v:int = int(min_r) if isinstance(min_r, int) else 0
                    _rowCount += (max_r_v - min_r_v + 1)
            
            _rowProcess: int = 0

            for tableName, tbl_info in _targetTables.items():
                _rowProcessTable:int=0
                sheetName = tbl_info['sheet']
                sheet = self.sourceWorkbook[sheetName]
                
                min_c, min_r, max_c, max_r = range_boundaries(tbl_info['ref'])
                if max_r is None: return 
                _dataTable = []
                _headers = []
                _isHead:bool=True
                if isinstance(sheet, (ReadOnlyWorksheet, Worksheet)):
                    for row in sheet.iter_rows(
                        min_row=min_r, max_row=max_r,
                        min_col=min_c, max_col=max_c,
                        values_only=True
                    ):
                        if _isHead: 
                            _headers = row
                            _isHead = False
                        else :
                            _dataTable.append(list(row))
                        _rowProcess += 1
                        _rowProcessTable += 1

                        if progressCallback and _rowCount > 0:
                            _value:float = float((_rowProcess / _rowCount) * 100)
                            _valueTable:float = float((_rowProcessTable / max_r) * 100)
                            progressCallback(
                                _value, 
                                ("Read -> \n")
                                    + (f"\t● Table: {tableName} ({sheetName}) \n")
                                    + (f"\t● (Row: {_rowProcess:05d} of {_rowCount:05d})"),
                                tableName,
                                _valueTable
                            )

                df = pd.DataFrame(_dataTable, columns=_headers)
                _selectColumns:list[str] | None = self.settingApp.GetColumnsByTable(tableName)
                if _selectColumns: 
                    df = df[_selectColumns]
                self.__addToDataFrame(tableName, df)

            if self.sourceWorkbookMeta:
                self.after(0, self.__wbLoadTableBySheet)
                self.after(0, lambda: self._cmdDataTable.config(state=tk.NORMAL))

        except Exception as e:
            self.after(0, self.__wbIssueLoadFile, str(e))

    def __wbProgressLoad(self, _value: float, _sheetData: str, _tableName:str, _tableProgress:float):
        if hasattr(self, '_excelProgressFile'): self._excelProgressFile.set(value=_value)
        self._lblStatus.config(text=f'Process: {_sheetData}')
        self.lstProgress.UpdateProgress(_tableName, _tableProgress)

    def __wbEndingLoad(self):
        self._lblStatus.config(text='Load successful')
        self._cmdOpen.config(state=tk.NORMAL)

    def __wbIssueLoadFile(self, _message: str):
        self._lblStatus.config(text=_message)
        self._cmdOpen.config(state=tk.NORMAL)

    def __wbLoadTableBySheet(self) -> list[str]:
        if not self.__isInSelf(self._cboTable): return []
        self.__setValuesComboBox(self._cboTable)
        if not self.__isInSelf(self.settingApp): return []
        values: list[str] = self.settingApp.GetTables()
        if values: self.__setValuesComboBox(self._cboTable, values)
        return values

    def __onTableSelected(self):
        table_name = self._excelTableActive.get()
        tree = self._twViewer

        if not table_name and not tree:
            return

        df = self.__getDataFrameByTable(table_name)
        for item in tree.get_children():
            tree.delete(item)

        if df is None: return
        self.hp.ControlTreeviewSetLayout(tree, df)
        
        df_clean = df.fillna('')

        for _, row in df_clean.iterrows():
            tree.insert('', tk.END, values=list(row))

    #endregion

    #region dataframe

    def __addToDataFrame(self, name:str, df:DataFrame):
        if self._dfDataList is None: self._dfDataList = []
        if self._dfDataListIsProcess is None: self._dfDataListIsProcess = []
        tableIsProcess:list[str] = self._lstChkSheetIsProcess.GetCheckedItems()
        if not tableIsProcess:
            self._dfDataList.append(DataTypeDataTable(name, df))
        if name in tableIsProcess: self._dfDataListIsProcess.append(DataTypeDataTable(name, df))
        else: self._dfDataList.append(DataTypeDataTable(name, df))

    def __getDataFrameByTable(self, name:str)->pd.DataFrame | None:
        if not self.__isInSelf(self._dfDataList): return None
        if not self.__isInSelf(self._dfDataListIsProcess): return None
        if not name or len(name) <= 0: return None
        _dfs:list[DataTypeDataTable] = self._dfDataListIsProcess if self._dfDataListIsProcess else []
        _dfs.extend(self._dfDataList if self._dfDataList else [])
        if not _dfs: return None
        try:
            return next((s.df for s in _dfs if s.name == name), None)
        except:
            return None

    #endregion

    

if __name__ == "__main__":
    app = ImportProcess(themename='superhero')
    app.mainloop()
